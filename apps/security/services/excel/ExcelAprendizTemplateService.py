from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from django.db import transaction
from django.contrib.auth.hashers import make_password
from io import BytesIO
from apps.general.entity.models.Center import Center
from apps.general.entity.models.Regional import Regional
from apps.general.entity.models.Sede import Sede
from apps.security.entity.models import Person, User
from apps.general.entity.models import Program, Ficha, Apprentice
from apps.security.entity.models.DocumentType import DocumentType
from apps.security.emails.SendEmailsActivate import enviar_activacion_usuario
import string
import random

class ExcelAprendizTemplateService:
    """
    Servicio para generar plantillas de Excel para el registro masivo de aprendices.
    """
    def __init__(self):
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.required_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.data_style = {
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _get_document_types(self):
        """Obtiene las siglas de tipos de documento desde la BD"""
        return list(DocumentType.objects.filter(active=True).values_list('acronyms', flat=True))

    def _get_document_type_display_values(self):
        """Obtiene pares (acronyms, name) de tipos de documento desde la BD"""
        return list(DocumentType.objects.filter(active=True).values_list('acronyms', 'name'))

    def _apply_style(self, cell, style_dict):
        for style_type, style_value in style_dict.items():
            setattr(cell, style_type, style_value)

    def _auto_adjust_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_data_validation(self, worksheet, column_letter, values, start_row=2, end_row=1000):
        if not values:
            return
        values_str = ','.join([str(v) for v in values])
        dv = DataValidation(
            type="list",
            formula1=f'"{values_str}"',
            allow_blank=True
        )
        dv.error = 'El valor debe ser seleccionado de la lista'
        dv.errorTitle = 'Valor Inválido'
        dv.prompt = 'Selecciona un valor de la lista desplegable'
        dv.promptTitle = 'Seleccionar Valor'
        dv.add(f'{column_letter}{start_row}:{column_letter}{end_row}')
        worksheet.add_data_validation(dv)

    def generate_aprendiz_template(self):
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Aprendices"
        headers = [
            ('TIPO IDENTIFICACIÓN*', True),
            ('NÚMERO IDENTIFICACIÓN*', True),
            ('PRIMER NOMBRE*', True),
            ('SEGUNDO NOMBRE', False),
            ('PRIMER APELLIDO*', True),
            ('SEGUNDO APELLIDO', False),
            ('CORREO INSTITUCIONAL*', True),
            ('NÚMERO DE CELULAR*', True),
        ]
        for col_idx, (header, is_required) in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            style = self.required_style if is_required else self.header_style
            self._apply_style(cell, style)
        example_data = [
            ['CC', '87654321', 'María', 'José', 'García', 'López', 
             'maria.garcia@soy.sena.edu.co', '3007654321']
        ]
        for row_idx, data_row in enumerate(example_data, 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, self.data_style)
        
        # Crear hojas auxiliares con datos de la BD
        self._create_identification_types_sheet(wb)
        
        # Agregar validaciones de datos (listas desplegables)
        self._add_aprendiz_data_validations(ws_main)
        
        # Crear hoja de instrucciones
        self._create_aprendiz_instructions_sheet(wb)
        
        # Ajustar columnas
        self._auto_adjust_columns(ws_main)
        
        return self._save_workbook_to_response(wb, 'plantilla_aprendices.xlsx')

    def _create_programs_sheet(self, workbook):
        """Crea una hoja con los programas disponibles"""
        ws = workbook.create_sheet("Programas")
        headers = ['CÓDIGO', 'NOMBRE', 'TIPO', 'DESCRIPCIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        programs = Program.objects.filter(active=True).order_by('codeProgram')
        for row_idx, program in enumerate(programs, 2):
            ws.cell(row=row_idx, column=1, value=program.codeProgram)
            ws.cell(row=row_idx, column=2, value=program.name)
            ws.cell(row=row_idx, column=3, value=program.typeProgram)
            ws.cell(row=row_idx, column=4, value=program.description)
        self._auto_adjust_columns(ws)

    def _create_fichas_sheet(self, workbook):
        """Crea una hoja con las fichas disponibles"""
        ws = workbook.create_sheet("Fichas")
        headers = ['NÚMERO FICHA', 'PROGRAMA', 'CÓDIGO PROGRAMA']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        fichas = Ficha.objects.filter(active=True).select_related('program').order_by('file_number')
        for row_idx, ficha in enumerate(fichas, 2):
            ws.cell(row=row_idx, column=1, value=ficha.file_number)
            ws.cell(row=row_idx, column=2, value=ficha.program.name)
            ws.cell(row=row_idx, column=3, value=ficha.program.codeProgram)
        self._auto_adjust_columns(ws)

    def _create_identification_types_sheet(self, workbook):
        """Crea una hoja con los tipos de identificación desde la BD"""
        ws = workbook.create_sheet("Tipos de Identificación")
        headers = ['ID', 'SIGLAS', 'NOMBRE']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        document_types = DocumentType.objects.filter(active=True).order_by('acronyms')
        for row_idx, doc_type in enumerate(document_types, 2):
            ws.cell(row=row_idx, column=1, value=doc_type.id)
            ws.cell(row=row_idx, column=2, value=doc_type.acronyms)
            ws.cell(row=row_idx, column=3, value=doc_type.name)
        self._auto_adjust_columns(ws)

    def _create_regionales_sheet(self, workbook):
        """Crea una hoja con las regionales disponibles"""
        ws = workbook.create_sheet("Regionales")
        headers = ['ID', 'NOMBRE']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        regionales = Regional.objects.filter(active=True).order_by('name')
        for row_idx, regional in enumerate(regionales, 2):
            ws.cell(row=row_idx, column=1, value=regional.id)
            ws.cell(row=row_idx, column=2, value=regional.name)
        self._auto_adjust_columns(ws)

    def _create_centros_formacion_sheet(self, workbook):
        """Crea una hoja con los centros de formación disponibles"""
        ws = workbook.create_sheet("Centros de Formación")
        headers = ['ID', 'NOMBRE', 'REGIONAL']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        centros = Center.objects.filter(active=True).select_related('regional').order_by('name')
        for row_idx, centro in enumerate(centros, 2):
            ws.cell(row=row_idx, column=1, value=centro.id)
            ws.cell(row=row_idx, column=2, value=centro.name)
            ws.cell(row=row_idx, column=3, value=centro.regional.name if centro.regional else '')
        self._auto_adjust_columns(ws)

    def _create_sedes_sheet(self, workbook):
        """Crea una hoja con las sedes disponibles"""
        ws = workbook.create_sheet("Sedes")
        headers = ['ID', 'NOMBRE', 'CENTRO DE FORMACIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        sedes = Sede.objects.filter(active=True).select_related('center').order_by('name')
        for row_idx, sede in enumerate(sedes, 2):
            ws.cell(row=row_idx, column=1, value=sede.id)
            ws.cell(row=row_idx, column=2, value=sede.name)
            ws.cell(row=row_idx, column=3, value=sede.center.name if sede.center else '')
        self._auto_adjust_columns(ws)

    def _create_aprendiz_instructions_sheet(self, workbook):
        """Crea una hoja con instrucciones para aprendices"""
        ws = workbook.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA EL REGISTRO MASIVO DE APRENDICES",
            "",
            "CAMPOS OBLIGATORIOS (marcados con *):",
            "• Tipo de Identificación: Usar lista desplegable",
            "• Número de Identificación: Documento único sin puntos ni espacios",
            "• Primer Nombre y Primer Apellido: Nombres completos",
            "• Correo Institucional: Debe terminar en @soy.sena.edu.co",
            "• Número de Celular: Número de contacto (10 dígitos)",
            "",
            "CAMPOS OPCIONALES:",
            "• Segundo Nombre y Segundo Apellido: Pueden quedar vacíos",
            "",
            "FORMATOS IMPORTANTES:",
            "• Email: usuario@soy.sena.edu.co",
            "• Teléfono: Solo números (ejemplo: 3001234567)",
            "• Documento: Solo números sin puntos (ejemplo: 1234567890)",
            "",
            "NOTAS:",
            "• No modificar los nombres de las columnas",
            "• Los campos opcionales pueden quedar vacíos",
            "• Usar las listas desplegables para evitar errores de escritura",
            "• La contraseña se generará automáticamente y se enviará por correo",
            "• Los aprendices deben usar el dominio @soy.sena.edu.co"
        ]
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                self._apply_style(cell, self.header_style)
            elif instruction.startswith("•"):
                cell.font = Font(italic=True)
        self._auto_adjust_columns(ws)

    def _add_aprendiz_data_validations(self, worksheet):
        """Agrega validaciones de datos (listas desplegables) para la plantilla de aprendices"""
        # Validación para Tipo de Identificación (columna A)
        id_types = self._get_document_types()
        self._add_data_validation(worksheet, 'A', id_types)

    def _save_workbook_to_response(self, workbook, filename):
        """Guarda el workbook en un HttpResponse para descarga"""
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _generate_additional_password_chars(self, length=2):
        """Genera caracteres adicionales aleatorios para la contraseña"""
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))

    def _send_credentials_email(self, email, first_name, last_name, password):
        """Envía un correo con las credenciales de acceso usando el módulo existente con reintentos"""
        import time
        max_retries = 3
        retry_delay = 2  # segundos entre reintentos
        base_delay = 1   # delay base entre envíos para no sobrecargar el servidor
        
        for attempt in range(max_retries):
            try:
                # Delay base para no sobrecargar el servidor de correo
                if attempt > 0:  # No delay en el primer intento
                    time.sleep(retry_delay)
                
                full_name = f"{first_name} {last_name}"
                enviar_activacion_usuario(
                    email_destino=email,
                    nombre=full_name,
                    email_usuario=email,
                    password_temporal=password
                )
                
                # Si llegamos aquí, el envío fue exitoso
                # Pequeño delay para no sobrecargar el servidor
                time.sleep(base_delay)
                print(f"✅ Correo enviado exitosamente a {email} (intento {attempt + 1})")
                return True
                
            except Exception as e:
                print(f"❌ Error enviando correo a {email} (intento {attempt + 1}): {e}")
                if attempt == max_retries - 1:  # Último intento
                    print(f"🚫 Falló envío de correo a {email} después de {max_retries} intentos")
                    return False
                else:
                    print(f"🔄 Reintentando envío de correo a {email} en {retry_delay} segundos...")
        
        return False

    def process_aprendiz_excel(self, excel_file):
        """
        Procesa un archivo Excel con datos de aprendices para registro masivo.
        Los usuarios creados quedan activos automáticamente y se envían credenciales por correo.
        """
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            results = {
                'success': [],
                'errors': [],
                'total_processed': 0,
                'successful_registrations': 0,
                'emails_sent': 0
            }
            
            # Obtener los datos a partir de la fila 2
            for row_num in range(2, worksheet.max_row + 1):
                try:
                    results['total_processed'] += 1
                    
                    # Leer datos de la fila según las nuevas columnas
                    row_data = {
                        'tipo_identificacion': self._get_cell_value(worksheet, row_num, 1),
                        'numero_identificacion': self._get_cell_value(worksheet, row_num, 2),
                        'primer_nombre': self._get_cell_value(worksheet, row_num, 3),
                        'segundo_nombre': self._get_cell_value(worksheet, row_num, 4),
                        'primer_apellido': self._get_cell_value(worksheet, row_num, 5),
                        'segundo_apellido': self._get_cell_value(worksheet, row_num, 6),
                        'email': self._get_cell_value(worksheet, row_num, 7),
                        'telefono': self._get_cell_value(worksheet, row_num, 8),
                    }
                    
                    # Validar que los campos obligatorios estén presentes
                    validation_errors = self._validate_aprendiz_data(row_data)
                    if validation_errors:
                        results['errors'].append({
                            'row': row_num,
                            'errors': validation_errors,
                            'data': row_data
                        })
                        continue
                    
                    # Generar contraseña con caracteres adicionales
                    additional_chars = self._generate_additional_password_chars()
                    final_password = str(row_data['numero_identificacion']) + additional_chars
                    
                    # Procesar el registro
                    with transaction.atomic():
                        user_created = self._create_aprendiz_record(row_data, final_password)
                        if user_created:
                            results['successful_registrations'] += 1
                            
                            # Enviar correo con credenciales
                            email_sent = self._send_credentials_email(
                                row_data['email'],
                                row_data['primer_nombre'],
                                row_data['primer_apellido'],
                                final_password
                            )
                            
                            if email_sent:
                                results['emails_sent'] += 1
                            
                            results['success'].append({
                                'row': row_num,
                                'message': f"Aprendiz {row_data['primer_nombre']} {row_data['primer_apellido']} registrado exitosamente",
                                'email': row_data['email'],
                                'email_sent': email_sent,
                                'password': final_password  # Solo para debugging, remover en producción
                            })
                        else:
                            results['errors'].append({
                                'row': row_num,
                                'errors': ['Error al crear el registro del aprendiz'],
                                'data': row_data
                            })
                
                except Exception as e:
                    results['errors'].append({
                        'row': row_num,
                        'errors': [f"Error procesando fila: {str(e)}"],
                        'data': row_data if 'row_data' in locals() else {}
                    })
            
            return results
            
        except Exception as e:
            return {
                'success': [],
                'errors': [{'general': f"Error procesando archivo: {str(e)}"}],
                'total_processed': 0,
                'successful_registrations': 0,
                'emails_sent': 0
            }

    def _get_cell_value(self, worksheet, row, column):
        """Obtiene el valor de una celda, manejando valores None"""
        cell = worksheet.cell(row=row, column=column)
        return cell.value if cell.value is not None else ''

    def _validate_aprendiz_data(self, data):
        """Valida los datos de un aprendiz"""
        errors = []
        
        # Validar campos obligatorios
        required_fields = [
            ('tipo_identificacion', 'Tipo de Identificación'),
            ('numero_identificacion', 'Número de Identificación'),
            ('primer_nombre', 'Primer Nombre'),
            ('primer_apellido', 'Primer Apellido'),
            ('email', 'Correo Institucional'),
            ('telefono', 'Número de Celular'),
        ]
        
        for field, name in required_fields:
            if not data.get(field) or str(data.get(field)).strip() == '':
                errors.append(f"{name} es obligatorio")
        
        # Validar email institucional
        email = data.get('email', '')
        if email and not email.endswith('@soy.sena.edu.co'):
            errors.append('El email debe ser institucional (@soy.sena.edu.co)')
        
        # Validar que el email no esté repetido
        if email and User.objects.filter(email=email).exists():
            errors.append('El email ya está registrado')
        
        # Validar que el número de identificación no esté repetido
        numero_id = data.get('numero_identificacion', '')
        if numero_id and Person.objects.filter(number_identification=numero_id).exists():
            errors.append('El número de identificación ya está registrado')
        
        return errors

    def _create_aprendiz_record(self, data, final_password):
        """Crea un registro completo de aprendiz (Person + User)"""
        try:
            # 1. Obtener el ID del tipo de documento desde la BD
            doc_type = DocumentType.objects.get(acronyms=data['tipo_identificacion'], active=True)
            
            # 2. Crear Person usando type_identification_id directamente
            person = Person.objects.create(
                first_name=data['primer_nombre'],
                second_name=data.get('segundo_nombre', ''),
                first_last_name=data['primer_apellido'],
                second_last_name=data.get('segundo_apellido', ''),
                phone_number=data['telefono'],
                type_identification_id=doc_type.id,  # Usar el ID directamente
                number_identification=data['numero_identificacion'],
                active=True
            )
            
            # 3. Crear User (activo automáticamente)
            hashed_password = make_password(final_password)
            user = User.objects.create(
                email=data['email'],
                password=hashed_password,
                person=person,
                is_active=True,  # Activo automáticamente
                role_id=2,  # Rol de Aprendiz
                registered=False  # No registrado, ya que se activa automáticamente
            )
            
            # 4. Crear Aprendiz (sin ficha específica por ahora)
            aprendiz = Apprentice.objects.create(
                person=person,
                ficha=None,  # Se asignará posteriormente
                active=True
            )
            
            return True
            
        except Exception as e:
            print(f"Error creando aprendiz: {e}")
            return False
