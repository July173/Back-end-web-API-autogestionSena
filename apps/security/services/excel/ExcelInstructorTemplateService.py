from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from django.db import transaction
from django.contrib.auth.hashers import make_password
from io import BytesIO
from apps.security.entity.models import Role, Person, User
from apps.general.entity.models import KnowledgeArea, Instructor
from apps.security.entity.models.DocumentType import DocumentType
from apps.security.emails.SendEmailsActivate import enviar_activacion_usuario
from datetime import datetime
import string
import random
import os

class ExcelInstructorTemplateService:
    """
    Servicio para generar plantillas de Excel para el registro masivo de instructores.
    """
    def __init__(self):
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.required_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.data_style = {
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _get_document_types(self):
        """Obtiene las siglas de tipos de documento desde la BD"""
        return list(DocumentType.objects.filter(active=True).values_list('acronyms', flat=True))

    def _get_document_type_display_values(self):
        """Obtiene pares (acronyms, name) de tipos de documento desde la BD"""
        return list(DocumentType.objects.filter(active=True).values_list('acronyms', 'name'))

    def _apply_style(self, cell, style_dict):
        for style_type, style_value in style_dict.items():
            setattr(cell, style_type, style_value)

    def _auto_adjust_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_data_validation(self, worksheet, column_letter, values, start_row=2, end_row=1000, sheet_name=None, col_aux=None):
        # Siempre pasar los valores extraídos de la BD
        if not values:
            return
        values_str = ','.join([str(v) for v in values])
        if len(values_str) < 250 and all(',' not in str(v) for v in values):
            dv = DataValidation(
                type="list",
                formula1=f'"{values_str}"',
                allow_blank=True
            )
        elif sheet_name and col_aux and len(values) > 0:
            # Si el nombre de la hoja tiene espacios, poner comillas simples
            if ' ' in sheet_name:
                sheet_name_formula = f"'{sheet_name}'"
            else:
                sheet_name_formula = sheet_name
            dv = DataValidation(
                type="list",
                formula1=f'={sheet_name_formula}!${col_aux}$2:${col_aux}${len(values)+1}',
                allow_blank=True
            )
        else:
            return
        dv.error = 'El valor debe ser seleccionado de la lista'
        dv.errorTitle = 'Valor Inválido'
        dv.prompt = 'Selecciona un valor de la lista desplegable'
        dv.promptTitle = 'Seleccionar Valor'
        dv.add(f'{column_letter}{start_row}:{column_letter}{end_row}')
        worksheet.add_data_validation(dv)

    def generate_instructor_template(self):
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Instructores"
        headers = [
            ('TIPO IDENTIFICACIÓN*', True),
            ('NÚMERO IDENTIFICACIÓN*', True),
            ('PRIMER NOMBRE*', True),
            ('SEGUNDO NOMBRE', False),
            ('PRIMER APELLIDO*', True),
            ('SEGUNDO APELLIDO', False),
            ('CORREO INSTITUCIONAL*', True),
            ('NÚMERO DE CELULAR*', True),
            ('REGIONAL*', True),
            ('CENTRO DE FORMACIÓN*', True),
            ('SEDE DE FORMACIÓN*', True),
            ('ÁREA DE CONOCIMIENTO*', True),
            ('TIPO DE CONTRATO*', True),
            ('FECHA INICIO CONTRATO*', True),
            ('FECHA DE TERMINACIÓN DE CONTRATO*', True)
        ]
        for col_idx, (header, is_required) in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            style = self.required_style if is_required else self.header_style
            self._apply_style(cell, style)
        example_data = [
            ['CC', '1023456789', 'Juan', 'Carlos', 'Pérez', 'Gómez', 'juan.perez@sena.edu.co', '3004567890',
             'Huila', 'Centro de la Biotecnología Agropecuaria', 'Sede Principal Biotecnología',
             'Diseño', 'Planta', '2024-01-15', '2024-12-31']
        ]
        for row_idx, data_row in enumerate(example_data, 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, self.data_style)
        
        # Crear hojas auxiliares con datos de la BD
        self._create_knowledge_areas_sheet(wb)
        self._create_contract_types_sheet(wb)
        self._create_identification_types_sheet(wb)
        self._create_regionales_sheet(wb)
        self._create_centros_formacion_sheet(wb)
        self._create_sedes_sheet(wb)
        
        # Agregar validaciones de datos (listas desplegables)
        self._add_instructor_data_validations(ws_main)
        
        # Crear hoja de instrucciones
        self._create_instructor_instructions_sheet(wb)
        
        # Ajustar columnas
        self._auto_adjust_columns(ws_main)
        
        return self._save_workbook_to_response(wb, 'plantilla_instructores.xlsx')

    def _create_knowledge_areas_sheet(self, workbook):
        """Crea una hoja con las áreas de conocimiento disponibles"""
        ws = workbook.create_sheet("Áreas de Conocimiento")
        headers = ['ID', 'NOMBRE', 'DESCRIPCIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        knowledge_areas = KnowledgeArea.objects.filter(active=True).order_by('name')
        for row_idx, area in enumerate(knowledge_areas, 2):
            ws.cell(row=row_idx, column=1, value=area.id)
            ws.cell(row=row_idx, column=2, value=area.name)
            ws.cell(row=row_idx, column=3, value=area.description or '')
        self._auto_adjust_columns(ws)

    def _create_contract_types_sheet(self, workbook):
        """Crea una hoja con los tipos de contrato disponibles desde la BD"""
        from apps.general.entity.models.TypeContract import TypeContract
        ws = workbook.create_sheet("Tipos de Contrato")
        headers = ['ID', 'NOMBRE', 'DESCRIPCIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        contract_types = TypeContract.objects.filter(active=True).order_by('name')
        for row_idx, contract_type in enumerate(contract_types, 2):
            ws.cell(row=row_idx, column=1, value=contract_type.id)
            ws.cell(row=row_idx, column=2, value=contract_type.name)
            ws.cell(row=row_idx, column=3, value=contract_type.description or '')
        self._auto_adjust_columns(ws)

    def _create_identification_types_sheet(self, workbook):
        """Crea una hoja con los tipos de identificación desde la BD"""
        ws = workbook.create_sheet("Tipos de Identificación")
        headers = ['ID', 'SIGLAS', 'NOMBRE']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        document_types = DocumentType.objects.filter(active=True).order_by('acronyms')
        for row_idx, doc_type in enumerate(document_types, 2):
            ws.cell(row=row_idx, column=1, value=doc_type.id)
            ws.cell(row=row_idx, column=2, value=doc_type.acronyms)
            ws.cell(row=row_idx, column=3, value=doc_type.name)
        self._auto_adjust_columns(ws)

    def _create_regionales_sheet(self, workbook):
        """Crea una hoja con las regionales disponibles"""
        from apps.general.entity.models import Regional
        ws = workbook.create_sheet("Regionales")
        headers = ['ID', 'NOMBRE']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        regionales = Regional.objects.filter(active=True).order_by('name')
        for row_idx, regional in enumerate(regionales, 2):
            ws.cell(row=row_idx, column=1, value=regional.id)
            ws.cell(row=row_idx, column=2, value=regional.name)
        self._auto_adjust_columns(ws)

    def _create_centros_formacion_sheet(self, workbook):
        """Crea una hoja con los centros de formación disponibles"""
        from apps.general.entity.models import Center
        ws = workbook.create_sheet("Centros de Formación")
        headers = ['ID', 'NOMBRE', 'REGIONAL']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        centros = Center.objects.filter(active=True).select_related('regional').order_by('name')
        for row_idx, centro in enumerate(centros, 2):
            ws.cell(row=row_idx, column=1, value=centro.id)
            ws.cell(row=row_idx, column=2, value=centro.name)
            ws.cell(row=row_idx, column=3, value=centro.regional.name if centro.regional else '')
        self._auto_adjust_columns(ws)

    def _create_sedes_sheet(self, workbook):
        """Crea una hoja con las sedes disponibles"""
        from apps.general.entity.models import Sede
        ws = workbook.create_sheet("Sedes")
        headers = ['ID', 'NOMBRE', 'CENTRO DE FORMACIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        sedes = Sede.objects.filter(active=True).select_related('center').order_by('name')
        for row_idx, sede in enumerate(sedes, 2):
            ws.cell(row=row_idx, column=1, value=sede.id)
            ws.cell(row=row_idx, column=2, value=sede.name)
            ws.cell(row=row_idx, column=3, value=sede.center.name if sede.center else '')
        self._auto_adjust_columns(ws)

    def _create_instructor_instructions_sheet(self, workbook):
        """Crea una hoja con instrucciones para instructores"""
        ws = workbook.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA EL REGISTRO MASIVO DE INSTRUCTORES",
            "",
            "CAMPOS OBLIGATORIOS (marcados con *):",
            "• Primer Nombre, Primer Apellido: Nombres completos",
            "• Tipo y Número de Identificación: Usar lista desplegable para tipo",
            "• Teléfono: Número de contacto (incluir código de país si es internacional)",
            "• Email Institucional: Debe terminar en @sena.edu.co",
            "• Tipo de Contrato: Seleccionar de la lista desplegable",
            "• Fechas de Contrato: Formato YYYY-MM-DD (Ej: 2024-01-15)",
            "• Área de Conocimiento: Seleccionar de la lista desplegable",
            "• Regional, Centro de Formación, Sede: Seleccionar de las listas desplegables",
            "",
            "LISTAS DESPLEGABLES DISPONIBLES:",
            "• Tipo de Identificación: Haz clic en la celda para ver opciones",
            "• Tipo de Contrato: Haz clic en la celda para ver opciones",
            "• Área de Conocimiento: Haz clic en la celda para ver opciones",
            "• Regional, Centro de Formación, Sede: Haz clic en las celdas para ver opciones",
            "",
            "FORMATOS IMPORTANTES:",
            "• Fechas: YYYY-MM-DD",
            "• Email: usuario@sena.edu.co",
            "• Teléfono: Solo números y signos + -",
            "",
            "NOTAS:",
            "• No modificar los nombres de las columnas",
            "• Los campos opcionales pueden quedar vacíos",
            "• Usar las listas desplegables para evitar errores de escritura",
            "• La contraseña se generará automáticamente y se enviará por correo"
        ]
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:
                self._apply_style(cell, self.header_style)
            elif instruction.startswith("•"):
                cell.font = Font(italic=True)
        self._auto_adjust_columns(ws)

    def _add_instructor_data_validations(self, worksheet):
        """Agrega validaciones de datos (listas desplegables) para la plantilla de instructores"""
        # Tipo de Identificación (columna A)
        id_types = self._get_document_types()
        self._add_data_validation(worksheet, 'A', id_types, sheet_name="Tipos de Identificación", col_aux="A")

        # Regional (columna I)
        from apps.general.entity.models import Regional
        regionales = list(Regional.objects.filter(active=True).values_list('name', flat=True))
        self._add_data_validation(worksheet, 'I', regionales, sheet_name="Regionales", col_aux="B")

        # Centro de Formación (columna J)
        from apps.general.entity.models import Center
        centros = list(Center.objects.filter(active=True).values_list('name', flat=True))
        self._add_data_validation(worksheet, 'J', centros, sheet_name="Centros de Formación", col_aux="B")

        # Sede de Formación (columna K)
        from apps.general.entity.models import Sede
        sedes = list(Sede.objects.filter(active=True).values_list('name', flat=True))
        self._add_data_validation(worksheet, 'K', sedes, sheet_name="Sedes", col_aux="B")

        # Área de Conocimiento (columna L)
        knowledge_areas = list(KnowledgeArea.objects.filter(active=True).values_list('name', flat=True))
        self._add_data_validation(worksheet, 'L', knowledge_areas, sheet_name="Áreas de Conocimiento", col_aux="B")

        # Tipo de Contrato (columna M)
        from apps.general.entity.models.TypeContract import TypeContract
        contract_types = list(TypeContract.objects.filter(active=True).values_list('name', flat=True))
        self._add_data_validation(worksheet, 'M', contract_types, sheet_name="Tipos de Contrato", col_aux="B")

    def _save_workbook_to_response(self, workbook, filename):
        """Guarda el workbook en un HttpResponse para descarga"""
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _generate_additional_password_chars(self, length=2):
        """Genera caracteres adicionales aleatorios para la contraseña"""
        characters = string.ascii_letters + string.digits
        return ''.join(random.choice(characters) for _ in range(length))

    def _send_credentials_email(self, email, first_name, last_name, password):
        """Envía un correo con las credenciales de acceso usando el módulo existente con reintentos"""
        import time
        max_retries = 3
        retry_delay = 2  # segundos entre reintentos
        base_delay = 1   # delay base entre envíos para no sobrecargar el servidor
        
        for attempt in range(max_retries):
            try:
                # Delay base para no sobrecargar el servidor de correo
                if attempt > 0:  # No delay en el primer intento
                    time.sleep(retry_delay)
                
                full_name = f"{first_name} {last_name}"
                enviar_activacion_usuario(
                    email_destino=email,
                    nombre=full_name,
                    email_usuario=email,
                    password_temporal=password
                )
                
                # Si llegamos aquí, el envío fue exitoso
                # Pequeño delay para no sobrecargar el servidor
                time.sleep(base_delay)
                print(f"✅ Correo enviado exitosamente a {email} (intento {attempt + 1})")
                return True
                
            except Exception as e:
                print(f"❌ Error enviando correo a {email} (intento {attempt + 1}): {e}")
                if attempt == max_retries - 1:  # Último intento
                    print(f"🚫 Falló envío de correo a {email} después de {max_retries} intentos")
                    return False
                else:
                    print(f"🔄 Reintentando envío de correo a {email} en {retry_delay} segundos...")
        
        return False

    def process_instructor_excel(self, excel_file):
        """
        Procesa un archivo Excel con datos de instructores para registro masivo.
        Los usuarios creados quedan activos automáticamente y se envían credenciales por correo.
        """
        try:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active
            
            results = {
                'success': [],
                'errors': [],
                'total_processed': 0,
                'successful_registrations': 0,
                'emails_sent': 0
            }
            
            # Obtener los datos a partir de la fila 2 (la 1 son headers)
            for row_num in range(2, worksheet.max_row + 1):
                try:
                    results['total_processed'] += 1
                    
                    # Leer datos de la fila (según la nueva estructura)
                    row_data = {
                        'tipo_identificacion': self._get_cell_value(worksheet, row_num, 1),
                        'numero_identificacion': self._get_cell_value(worksheet, row_num, 2),
                        'primer_nombre': self._get_cell_value(worksheet, row_num, 3),
                        'segundo_nombre': self._get_cell_value(worksheet, row_num, 4),
                        'primer_apellido': self._get_cell_value(worksheet, row_num, 5),
                        'segundo_apellido': self._get_cell_value(worksheet, row_num, 6),
                        'email': self._get_cell_value(worksheet, row_num, 7),
                        'telefono': self._get_cell_value(worksheet, row_num, 8),
                        'regional': self._get_cell_value(worksheet, row_num, 9),
                        'centro_formacion': self._get_cell_value(worksheet, row_num, 10),
                        'sede': self._get_cell_value(worksheet, row_num, 11),
                        'area_conocimiento': self._get_cell_value(worksheet, row_num, 12),
                        'tipo_contrato': self._get_cell_value(worksheet, row_num, 13),
                        'fecha_inicio': self._get_cell_value(worksheet, row_num, 14),
                        'fecha_fin': self._get_cell_value(worksheet, row_num, 15),
                    }
                    
                    # Validar que los campos obligatorios estén presentes
                    validation_errors = self._validate_instructor_data(row_data)
                    if validation_errors:
                        results['errors'].append({
                            'row': row_num,
                            'errors': validation_errors,
                            'data': row_data
                        })
                        continue
                    
                    # Generar contraseña con caracteres adicionales
                    additional_chars = self._generate_additional_password_chars()
                    final_password = str(row_data['numero_identificacion']) + additional_chars
                    
                    # Procesar el registro
                    with transaction.atomic():
                        user_created = self._create_instructor_record(row_data, final_password)
                        if user_created:
                            results['successful_registrations'] += 1
                            
                            # Enviar correo con credenciales
                            email_sent = self._send_credentials_email(
                                row_data['email'],
                                row_data['primer_nombre'],
                                row_data['primer_apellido'],
                                final_password
                            )
                            
                            if email_sent:
                                results['emails_sent'] += 1
                            
                            results['success'].append({
                                'row': row_num,
                                'message': f"Instructor {row_data['primer_nombre']} {row_data['primer_apellido']} registrado exitosamente",
                                'email': row_data['email'],
                                'email_sent': email_sent,
                                'password': final_password  # Solo para debugging, remover en producción
                            })
                        else:
                            results['errors'].append({
                                'row': row_num,
                                'errors': ['Error al crear el registro del instructor'],
                                'data': row_data
                            })
                
                except Exception as e:
                    results['errors'].append({
                        'row': row_num,
                        'errors': [f"Error procesando fila: {str(e)}"],
                        'data': row_data if 'row_data' in locals() else {}
                    })
            
            return results
            
        except Exception as e:
            return {
                'success': [],
                'errors': [{'general': f"Error procesando archivo: {str(e)}"}],
                'total_processed': 0,
                'successful_registrations': 0,
                'emails_sent': 0
            }

    def _get_cell_value(self, worksheet, row, column):
        """Obtiene el valor de una celda, manejando valores None"""
        cell = worksheet.cell(row=row, column=column)
        return cell.value if cell.value is not None else ''

    def _validate_instructor_data(self, data):
        """Valida los datos de un instructor"""
        errors = []
        
        # Validar campos obligatorios
        required_fields = [
            ('tipo_identificacion', 'Tipo de Identificación'),
            ('numero_identificacion', 'Número de Identificación'),
            ('primer_nombre', 'Primer Nombre'),
            ('primer_apellido', 'Primer Apellido'), 
            ('email', 'Correo Institucional'),
            ('telefono', 'Número de Celular'),
            ('area_conocimiento', 'Área de Conocimiento'),
            ('tipo_contrato', 'Tipo de Contrato'),
            ('fecha_inicio', 'Fecha Inicio Contrato'),
            ('fecha_fin', 'Fecha de Terminación de Contrato'),
            ('regional', 'Regional'),
            ('centro_formacion', 'Centro de Formación'),
            ('sede', 'Sede de Formación')
        ]
        
        for field, name in required_fields:
            if not data.get(field) or str(data.get(field)).strip() == '':
                errors.append(f"{name} es obligatorio")
        
        # Validar email institucional
        email = data.get('email', '')
        if email and not email.endswith('@sena.edu.co'):
            errors.append('El email debe ser institucional (@sena.edu.co)')
        
        # Validar que el email no esté repetido
        if email and User.objects.filter(email=email).exists():
            errors.append('El email ya está registrado')
        
        # Validar que el número de identificación no esté repetido
        numero_id = data.get('numero_identificacion', '')
        if numero_id and Person.objects.filter(number_identification=numero_id).exists():
            errors.append('El número de identificación ya está registrado')
        
        # Validar que el área de conocimiento exista
        area_nombre = data.get('area_conocimiento', '')
        if area_nombre and not KnowledgeArea.objects.filter(name=area_nombre, active=True).exists():
            errors.append('El área de conocimiento no existe o no está activa')
        
        # Validar que el tipo de contrato exista en la BD
        from apps.general.entity.models.TypeContract import TypeContract
        tipo_contrato = data.get('tipo_contrato', '')
        if tipo_contrato and not TypeContract.objects.filter(name=tipo_contrato, active=True).exists():
            errors.append('El tipo de contrato no existe o no está activo')
        
        return errors

    def _create_instructor_record(self, data, final_password):
        """Crea un registro completo de instructor (Person + User + Instructor)"""
        try:
            from apps.general.entity.models.TypeContract import TypeContract
            
            # 1. Obtener el ID del tipo de documento desde la BD
            doc_type = DocumentType.objects.get(acronyms=data['tipo_identificacion'], active=True)
            
            # 2. Crear Person usando type_identification_id directamente
            person = Person.objects.create(
                first_name=data['primer_nombre'],
                second_name=data.get('segundo_nombre', ''),
                first_last_name=data['primer_apellido'],
                second_last_name=data.get('segundo_apellido', ''),
                phone_number=data['telefono'],
                type_identification_id=doc_type.id,  # Usar el ID directamente
                number_identification=data['numero_identificacion'],
                active=True
            )
            
            # 3. Crear User (activo automáticamente)
            hashed_password = make_password(final_password)
            user = User.objects.create(
                email=data['email'],
                password=hashed_password,
                person=person,
                is_active=True,  # Activo automáticamente
                role_id=3,  # Rol de Instructor
                registered=False  # No registrado
            )
            
            # 4. Obtener área de conocimiento
            knowledge_area = KnowledgeArea.objects.get(name=data['area_conocimiento'], active=True)
            
            # 5. Obtener tipo de contrato desde la BD
            contract_type = TypeContract.objects.get(name=data['tipo_contrato'], active=True)
            
            # 6. Crear Instructor
            instructor = Instructor.objects.create(
                person=person,
                contractType_id=contract_type.id,  # Usar el ID directamente
                contractStartDate=data['fecha_inicio'],
                contractEndDate=data['fecha_fin'],
                knowledgeArea=knowledge_area,
                active=True
            )
            
            return True
            
        except Exception as e:
            print(f"Error creando instructor: {e}")
            return False
